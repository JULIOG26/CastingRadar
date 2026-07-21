from pathlib import Path
from datetime import datetime
import shutil

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


class ExcelExporter:

    CABECERAS = [
        "Puntos",
        "Publicación",
        "Límite",
        "Interés",
        "Papel",
        "Producción",
        "Remunerado",
        "Derechos",
        "Ubicación",
        "Sexo",
        "Edad",
        "Título",
        "Empresa",
        "Ciudad",
        "Fuente",
        "URL",
        "Decisión",
        "Motivo",
        "Comentarios",
        "Revisado",
        "Presentado",
        "Resultado",
    ]

    def __init__(self):

        self.export_dir = Path("exports")
        self.export_dir.mkdir(exist_ok=True)

        self.backup_dir = self.export_dir / "backups"
        self.backup_dir.mkdir(exist_ok=True)

    def export(self, castings):

        excel = self.export_dir / "CastingRadar.xlsx"

        if excel.exists():

            backup = self.backup_dir / (
                f"CastingRadar_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
            )

            shutil.copy2(excel, backup)

            wb = load_workbook(excel)

            # Obtener o crear hoja ACTIVO
            if "ACTIVO" in wb.sheetnames:
                ws = wb["ACTIVO"]
            else:
                ws = wb.create_sheet("ACTIVO", 0)
                ws.append(self.CABECERAS)
                for c in ws[1]:
                    c.font = Font(bold=True)

            # Obtener o crear hoja HISTORICO
            if "HISTORICO" in wb.sheetnames:
                historico = wb["HISTORICO"]
            else:
                historico = wb.create_sheet("HISTORICO")
                historico.append(self.CABECERAS)
                for c in historico[1]:
                    c.font = Font(bold=True)

            ws = wb.active

        else:

            wb = Workbook()

            ws = wb.active

            ws.title = "CastingRadar"

            ws.append(self.CABECERAS)

            for c in ws[1]:
                c.font = Font(bold=True)

        urls = {}

        for fila in range(2, ws.max_row + 1):

            url = ws.cell(row=fila, column=16).value

            if url:
                urls[url] = fila

        for casting in castings:

            analisis = casting.analisis or {}

            edad = ""

            if (
                analisis.get("edad_min") is not None
                or analisis.get("edad_max") is not None
            ):
                edad = (
                    f"{analisis.get('edad_min','')}-"
                    f"{analisis.get('edad_max','')}"
                )

            puntos = casting.puntuacion

            if puntos >= 100:
                interes = "★★★★★"
            elif puntos >= 80:
                interes = "★★★★"
            elif puntos >= 60:
                interes = "★★★"
            elif puntos >= 40:
                interes = "★★"
            else:
                interes = "★"

            datos = [
                puntos,
                casting.fecha_publicacion,
                casting.fecha_limite,
                interes,
                analisis.get("papel", ""),
                analisis.get("produccion", ""),
                analisis.get("remunerado", ""),
                analisis.get("derechos", ""),
                analisis.get("ubicacion", ""),
                analisis.get("sexo", ""),
                edad,
                casting.titulo,
                casting.empresa,
                casting.ciudad,
                casting.fuente,
                casting.url,
            ]

            if casting.url in urls:

                fila = urls[casting.url]

                for col, valor in enumerate(datos, start=1):

                    ws.cell(row=fila, column=col).value = valor

            else:
                ws.append(
                    datos
                    + [
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                    ]
                )

                urls[casting.url] = ws.max_row

        for columna in ws.columns:

            longitud = 0

            letra = columna[0].column_letter

            for celda in columna:

                try:

                    valor = str(celda.value)

                except Exception:

                    valor = ""

                if len(valor) > longitud:

                    longitud = len(valor)

            ws.column_dimensions[letra].width = min(longitud + 2, 60)

        wb.save(excel)

        return excel